#!/usr/bin/env python3
"""
Create CPSY 1950 Final Project Grading Spreadsheet
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Grades"

# Define colors
HEADER_FILL = PatternFill(start_color="1E2761", end_color="1E2761", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFD9B2", end_color="FFD9B2", fill_type="solid")
REGULAR_FONT = Font(name="Arial", size=10)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = [
    "Group #",
    "Students",
    "Project Title",
    "Research Quality (/10)",
    "Use of LLMs (/10)",
    "Poster Quality (/5)",
    "Oral Presentation (/5)",
    "Depth & Rigor (/5)",
    "TOTAL (/35)",
    "Late?",
    "Comments"
]

# Add headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

# Grading data
grades_data = [
    # TOP TIER
    (1, "Haya Bugshan & Mohammad Essa", "VLM Feedforward vs Feedback", 10, 10, 5, 5, 5, "", False, "Excellent design, Marr framework, biological grounding"),
    (2, "Tommy Nguyen, Zhafira Fawnia, Shreya Gulati", "Loss Insensitivity / Iowa Gambling Task", 10, 10, 5, 5, 5, "", False, "Outstanding methodology, fine-tuning insights"),
    (3, "Edward Wibowo & Taarini Singh", "Incentive Effects Two-Armed Bandit", 10, 9, 5, 5, 5, "", False, "Strong confound controls, System 1/2 integration"),
    
    # STRONG
    (4, "Quintessa Frisch & Faizan Zaidi", "Cognitive Maps & Spatial Relations", 9, 9, 4, 5, 5, "", False, "Good bio comparisons, H2 supported"),
    (5, "Nicolas Kim & Amara Fuchs", "Odd-One-Out THINGS", 9, 8, 5, 5, 5, "", False, "Good interpretability, limited trial count"),
    (6, "Michelsen & Gomez", "VPT Transformed Images", 9, 9, 5, 5, 5, "", False, "Excellent counterfactual design"),
    (7, "Isaac Bitran & Varit Asavathiratham", "Odd-One-Out Multimodal", 8, 9, 5, 5, 5, "", False, "Good contamination control"),
    (8, "David Vassalluzzo & Shivashish Das", "Temporal Discounting", 9, 8, 5, 5, 5, "", False, "Good Marr integration, System 1/2"),
    (9, "Mindy Kim & Avni Rajpal", "THINGS Odd-One-Out Conceptual Similarity", 9, 9, 5, 5, 5, "", False, "Excellent scholarly work"),
    (10, "Daniel Demessie, Sydney Riccio, Maria Kim", "Amodal Completion", 9, 9, 5, 5, 5, "", False, "Strong confound controls"),
    (11, "Aura Machanic & Frances Robertson", "Language Framing Columbia Card Task", 9, 8, 5, 5, 5, "", False, "Clean manipulation, strong confound"),
    (12, "Aaryaa Kamdar & Josef Hatcher", "Hyperbolic Discounting", 9, 9, 5, 5, 5, "", False, "Exceptional theory (Mahowald, Marr, Centaur)"),
    (13, "Yoo Jin Lee & Vivian Wang", "Exploration-Exploitation CoT", 9, 9, 4, 5, 5, "", False, "Excellent negative result, algorithmic insight"),
    (14, "Ziqi Shen, Emmie Fitz-Gibbon, Sanskriti Manoharan", "Gabor Patch VLM", 9, 8, 5, 5, 5, "", False, "Comprehensive parametric testing"),
    (15, "Robayet Hossain & Aman Bhutani", "Iowa Gambling Risk Learning", 8, 9, 4, 5, 5, "", False, "Good trial-by-trial pipeline"),
    (16, "Athulith Paraselli & David Lin", "Visual vs Textual Category Learning", 8, 8, 5, 5, 5, "", False, "Innovative cross-modality"),
    (17, "Jinting He & Wanjing Zhou", "Framing Effect Memorization vs Generalization", 9, 8, 4, 5, 5, "", False, "Well-designed counterfactual, only 2 models"),
    (18, "Kenneth Cho & Coco Zeng", "Iowa Gambling Task", 8, 8, 5, 5, 5, "", False, "Good human baseline (N=617)"),
    (19, "Kamyar Mirfakhraie & Joseph Ricciardulli", "Age-related VPT", 8, 8, 5, 5, 5, "", False, "Novel linguistic manipulation"),
    (20, "Daniel Lee & Navya Sahay", "Medin Categorization VLM", 8, 8, 5, 5, 5, "", False, "Novel modality testing, only 1 model"),
    
    # GOOD
    (21, "Tilika Kulkarni & Ashley Park", "BART Risk Task", 8, 8, 4, 5, 5, "", False, "Multiple metrics, small human sample (N=5)"),
    (22, "Jeonghyeok Park & Robert Desmond", "Optimism Bias Casino Bandit", 8, 8, 4, 5, 5, "", True, "Centaur benchmark; Jeonghyeok: Dean's Note for special circumstances"),
    (23, "Dasha Dmitrieva & Siddhant Karmali", "Risky Choice Gambling", 8, 7, 4, 5, 5, "", True, "Good prediction vs understanding framing"),
    (24, "Josh Gerber, Alejandra Hernandez Moyers, Alex Sayette", "In-Context Learning Columbia Card Task", 9, 8, 4, 5, 5, "", True, "Strong methodology, good frontier comparison"),
    
    # NEEDS IMPROVEMENT
    (25, "Priscilla Maalouf & Bahar Charyyeva", "Logical Reasoning", 7, 7, 4, 5, 5, "", False, "Small sample, novel tasks (no benchmark)"),
]

# Add data rows with formulas
for row_idx, (group_num, students, title, rq, llm, poster, oral, depth, _, late, comments) in enumerate(grades_data, 2):
    ws.cell(row=row_idx, column=1).value = group_num
    ws.cell(row=row_idx, column=2).value = students
    ws.cell(row=row_idx, column=3).value = title
    ws.cell(row=row_idx, column=4).value = rq
    ws.cell(row=row_idx, column=5).value = llm
    ws.cell(row=row_idx, column=6).value = poster
    ws.cell(row=row_idx, column=7).value = oral
    ws.cell(row=row_idx, column=8).value = depth
    
    # Total formula: SUM of columns D-H
    total_cell = ws.cell(row=row_idx, column=9)
    total_cell.value = f"=SUM(D{row_idx}:H{row_idx})"
    
    ws.cell(row=row_idx, column=10).value = "Yes" if late else ""
    ws.cell(row=row_idx, column=11).value = comments
    
    # Apply formatting to all cells in row
    for col in range(1, 12):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = REGULAR_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Center alignment for numeric columns
        if col in [1, 4, 5, 6, 7, 8, 9, 10]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply conditional formatting on TOTAL column (column I)
    total_value_cell = ws.cell(row=row_idx, column=9)
    # Green for 33-35
    if row_idx in range(2, 5):  # Top tier groups (1-3)
        total_value_cell.fill = GREEN_FILL
    # Yellow for 30-32
    elif row_idx in range(5, 21):  # Strong groups (4-20)
        total_value_cell.fill = YELLOW_FILL
    # Light orange for 28-29
    elif row_idx in range(21, 26):  # Good and needs improvement groups
        if row_idx == 26:  # Group 25
            total_value_cell.fill = ORANGE_FILL
        else:
            total_value_cell.fill = YELLOW_FILL

# After all data, add notes row
notes_row = len(grades_data) + 3
ws.cell(row=notes_row, column=1).value = "NOTES:"
ws.cell(row=notes_row, column=1).font = Font(bold=True, name="Arial", size=10)

notes_content = (
    "Duplicate submission: Zaidi/Frisch (both members submitted). "
    "3 late submissions flagged (Groups 22, 23, 24). "
    "Oral = 5/5 for all per instructor."
)
ws.cell(row=notes_row, column=2).value = notes_content
ws.cell(row=notes_row, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Set column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 14
ws.column_dimensions['H'].width = 14
ws.column_dimensions['I'].width = 12
ws.column_dimensions['J'].width = 8
ws.column_dimensions['K'].width = 45

# Freeze header row
ws.freeze_panes = "A2"

# Save file
output_path = "/Users/tserre/Work/personal/website/cpsy1950/CPSY1950_Project_Grades.xlsx"

wb.save(output_path)
print(f"Spreadsheet created successfully: {output_path}")
